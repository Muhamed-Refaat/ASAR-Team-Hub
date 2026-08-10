import openpyxl
import csv
import os

def sync_database():
    excel_path = os.path.join('spreedsheet-copy', 'ASAR-Team-Hub.xlsx')
    dest_dir = 'spreedsheet-copy'

    if not os.path.exists(excel_path):
        print(f"Error: Master workbook '{excel_path}' not found.")
        return

    # Load master workbook
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    print(f"Parsing sheets from '{excel_path}'...")

    for sheetname in wb.sheetnames:
        sh = wb[sheetname]
        csv_path = os.path.join(dest_dir, f"{sheetname}.csv")
        
        # Extract headers (first non-empty column set)
        headers = []
        for c in range(1, sh.max_column+1):
            v = sh.cell(1, c).value
            if v is not None:
                headers.append(v)
            else:
                break
                
        if not headers:
            continue
            
        rows_written = 0
        with open(csv_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            
            for r in range(2, sh.max_row+1):
                row_vals = [sh.cell(r, c).value for c in range(1, len(headers)+1)]
                # Write row if at least one cell has content
                if any(v is not None and v != '' for v in row_vals):
                    row_serialized = []
                    for val in row_vals:
                        if hasattr(val, 'strftime'):
                            # Serialize dates to standard YYYY-MM-DD format
                            row_serialized.append(val.strftime('%Y-%m-%d'))
                        elif val is None:
                            row_serialized.append('')
                        else:
                            row_serialized.append(str(val))
                    writer.writerow(row_serialized)
                    rows_written += 1
                    
        print(f"  Created '{sheetname}.csv' -> {rows_written} records.")
        
    print("Local database synchronization complete!")

if __name__ == '__main__':
    sync_database()
