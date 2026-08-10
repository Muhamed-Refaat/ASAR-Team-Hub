import openpyxl

wb = openpyxl.load_workbook('spreedsheet-copy/ASAR-Team-Hub.xlsx')

with open('xlsx_structure.txt', 'w', encoding='utf-8') as f:
    f.write(f"Sheet Names Raw: {repr(wb.sheetnames)}\n")
    for name in wb.sheetnames:
        sh = wb[name]
        headers = [sh.cell(1, c).value for c in range(1, sh.max_column+1)]
        headers_repr = [h.replace('\n', '\\n') if isinstance(h, str) else str(h) for h in headers if h is not None]
        f.write(f"Sheet {repr(name)}: {headers_repr}\n")
print("Structure written to xlsx_structure.txt")
